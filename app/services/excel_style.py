from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)


class ExcelStyle:

    HEADER_FILL = PatternFill(
        fill_type="solid",
        fgColor="305496",
    )

    TITLE_FILL = PatternFill(
        fill_type="solid",
        fgColor="D9EAD3",
    )

    SUCCESS_FILL = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE",
    )

    ERROR_FILL = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE",
    )

    HEADER_FONT = Font(
        bold=True,
        color="FFFFFF",
        size=12,
    )

    TITLE_FONT = Font(
        bold=True,
        size=16,
    )

    DEFAULT_FONT = Font(
        size=11,
    )

    CENTER = Alignment(
        horizontal="center",
        vertical="center",
    )

    LEFT = Alignment(
        vertical="center",
    )

    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
