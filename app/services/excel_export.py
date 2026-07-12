from datetime import date
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.services.report_generator import ReportGeneratorService
from app.services.excel_style import ExcelStyle

from openpyxl.drawing.image import Image

from app.core.bot import bot
from app.services.photo_export import PhotoExportService


class ExcelExportService:
    """
    Excel PRO report generator.
    """

    def __init__(
        self,
        session: AsyncSession,
    ):
        self.session = session

        self.generator = ReportGeneratorService(
            session,
        )

        self.photo_export = PhotoExportService(
            bot,
        )

        self.output_dir = Path(
            "exports/reports"
        )

        self.output_dir.mkdir(
            parents=True,
            exist_ok=True,
        )

    # =====================================================
    # CREATE REPORT
    # =====================================================

    async def create_report(
        self,
        report_date: date | None = None,
    ) -> Path:
        """
        Create Excel report.
        """

        if report_date is None:
            report_date = date.today()

        dashboard_data = await self.generator.dashboard(
            report_date,
        )

        answers = await self.generator.answers(
            report_date,
        )

        photos = await self.generator.photos(
            report_date,
        )

        workbook = Workbook()

        dashboard_ws = workbook.active
        dashboard_ws.title = "Dashboard"

        answers_ws = workbook.create_sheet(
            "Ответы",
        )

        food_ws = workbook.create_sheet(
            "Питание",
        )

        training_ws = workbook.create_sheet(
            "Тренировки",
        )

        photos_ws = workbook.create_sheet(
            "Фото",
        )

        statistics_ws = workbook.create_sheet(
            "Статистика",
        )

        # Dashboard
        self._fill_dashboard(
            dashboard_ws,
            dashboard_data,
        )

        # Answers
        self._fill_answers(
            answers_ws,
            answers,
        )

        # Food
        self._fill_food(
            food_ws,
            answers,
        )

        # Training
        self._fill_training(
            training_ws,
            answers,
        )

        # Photos
        self._fill_photos(
            photos_ws,
            photos,
        )

        # Statistics
        self._fill_statistics(
            statistics_ws,
            dashboard_data,
            answers,
        )

        # Apply style
        for sheet in workbook.worksheets:

            self._apply_table_style(
                sheet,
            )

            self._auto_width(
                sheet,
            )

        filename = (
            f"Fitness_Report_{report_date}.xlsx"
        )

        filepath = self.output_dir / filename

        workbook.save(
            filepath,
        )

        return filepath

    # =====================================================
    # DASHBOARD
    # =====================================================

    def _fill_dashboard(
        self,
        ws,
        data,
    ):
        """
        Fill dashboard sheet.
        """

        ws.merge_cells("A1:D2")

        title = ws["A1"]

        title.value = "FITNESS CONTROLLER PRO"

        title.fill = ExcelStyle.HEADER_FILL
        title.font = ExcelStyle.TITLE_FONT
        title.alignment = ExcelStyle.CENTER

        ws["A4"] = "Дата формирования"
        ws["B4"] = data["date"].strftime(
            "%d.%m.%Y",
        )

        rows = [
            (
                "👥 Пользователей",
                data["total_users"],
            ),
            (
                "✅ Прошли",
                data["completed"],
            ),
            (
                "❌ Не прошли",
                data["not_completed"],
            ),
            (
                "📈 Выполнение",
                f"{data['percent']} %",
            ),
            (
                "📷 Фото",
                data["photos"],
            ),
            (
                "⚖ Средний вес",
                f"{data['average_weight']} кг",
            ),
        ]

        row = 6

        for title, value in rows:

            ws.cell(
                row=row,
                column=1,
                value=title,
            )

            ws.cell(
                row=row,
                column=2,
                value=value,
            )

            row += 1

        ws.freeze_panes = "A6"

    # =====================================================
    # COMMON STYLE
    # =====================================================

    def _apply_table_style(
        self,
        ws,
    ):
        """
        Apply common style.
        """

        for row in ws.iter_rows():

            for cell in row:

                if cell.value is None:
                    continue

                cell.border = ExcelStyle.BORDER

                if cell.row == 1:

                    cell.fill = ExcelStyle.HEADER_FILL

                    cell.font = ExcelStyle.HEADER_FONT

                    cell.alignment = ExcelStyle.CENTER

                else:

                    cell.font = ExcelStyle.DEFAULT_FONT

                    cell.alignment = ExcelStyle.LEFT

    # =====================================================
    # AUTO WIDTH
    # =====================================================

    def _auto_width(
        self,
        ws,
    ):
        """
        Auto width columns.
        """

        from openpyxl.utils import get_column_letter

        for column_cells in ws.columns:

            length = 0

            column = get_column_letter(
                column_cells[0].column,
            )

            for cell in column_cells:

                if cell.value is None:
                    continue

                value = str(
                    cell.value,
                )

                if len(value) > length:

                    length = len(
                        value,
                    )

            ws.column_dimensions[
                column
            ].width = min(
                length + 3,
                45,
            )

    # =====================================================
    # ANSWERS
    # =====================================================

    def _fill_answers(
        self,
        ws,
        rows,
    ):
        """
        Fill answers worksheet.
        """

        headers = [
            "Пользователь",
            "Дата",
            "Вопрос",
            "Ответ",
            "Фото",
            "Статус",
        ]

        for col, title in enumerate(
            headers,
            start=1,
        ):

            cell = ws.cell(
                row=1,
                column=col,
                value=title,
            )

            cell.fill = ExcelStyle.HEADER_FILL
            cell.font = ExcelStyle.HEADER_FONT
            cell.alignment = ExcelStyle.CENTER
            cell.border = ExcelStyle.BORDER

        row = 2

        for item in rows:

            ws.cell(
                row=row,
                column=1,
                value=item["user"],
            )

            ws.cell(
                row=row,
                column=2,
                value=item["date"],
            )

            ws.cell(
                row=row,
                column=3,
                value=item["question"],
            )

            ws.cell(
                row=row,
                column=4,
                value=item["answer"],
            )

            ws.cell(
                row=row,
                column=5,
                value=item["photos"],
            )

            status = ws.cell(
                row=row,
                column=6,
            )

            if item["photos"] > 0:

                status.value = "✅ Есть"

                status.fill = ExcelStyle.SUCCESS_FILL

            else:

                status.value = "❌ Нет"

                status.fill = ExcelStyle.ERROR_FILL

            for col in range(
                1,
                7,
            ):

                current = ws.cell(
                    row=row,
                    column=col,
                )

                current.border = ExcelStyle.BORDER

                if col != 6:

                    current.alignment = ExcelStyle.LEFT

            row += 1

        ws.freeze_panes = "A2"

        ws.auto_filter.ref = ws.dimensions

        widths = {
            "A": 28,
            "B": 15,
            "C": 45,
            "D": 45,
            "E": 10,
            "F": 16,
        }

        for column, width in widths.items():

            ws.column_dimensions[
                column
            ].width = width

    # =====================================================
    # FOOD
    # =====================================================

    def _fill_food(
        self,
        ws,
        rows,
    ):
        """
        Fill food worksheet.
        """

        headers = [
            "Пользователь",
            "Дата",
            "Приём пищи",
            "Ответ",
            "Фото",
        ]

        for col, title in enumerate(headers, start=1):

            cell = ws.cell(
                row=1,
                column=col,
                value=title,
            )

            cell.fill = ExcelStyle.HEADER_FILL
            cell.font = ExcelStyle.HEADER_FONT
            cell.alignment = ExcelStyle.CENTER
            cell.border = ExcelStyle.BORDER

        row = 2

        keywords = (
            "завтрак",
            "обед",
            "ужин",
            "нонушта",
            "тушлик",
            "кечки",
        )

        for item in rows:

            question = item["question"].lower()

            if not any(
                key in question
                for key in keywords
            ):
                continue

            ws.cell(row=row, column=1).value = item["user"]
            ws.cell(row=row, column=2).value = item["date"]
            ws.cell(row=row, column=3).value = item["question"]
            ws.cell(row=row, column=4).value = item["answer"]
            ws.cell(row=row, column=5).value = item["photos"]

            row += 1

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    # =====================================================
    # TRAINING
    # =====================================================

    def _fill_training(
        self,
        ws,
        rows,
    ):
        """
        Fill training worksheet.
        """

        headers = [
            "Пользователь",
            "Дата",
            "Показатель",
            "Ответ",
        ]

        for col, title in enumerate(headers, start=1):

            cell = ws.cell(
                row=1,
                column=col,
                value=title,
            )

            cell.fill = ExcelStyle.HEADER_FILL
            cell.font = ExcelStyle.HEADER_FONT
            cell.alignment = ExcelStyle.CENTER
            cell.border = ExcelStyle.BORDER

        row = 2

        keywords = (
            "вода",
            "сон",
            "шаг",
            "трен",
            "машк",
            "сув",
            "ухлаш",
        )

        for item in rows:

            question = item["question"].lower()

            if not any(
                key in question
                for key in keywords
            ):
                continue

            ws.cell(row=row, column=1).value = item["user"]
            ws.cell(row=row, column=2).value = item["date"]
            ws.cell(row=row, column=3).value = item["question"]
            ws.cell(row=row, column=4).value = item["answer"]

            row += 1

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    # =====================================================
    # STATISTICS
    # =====================================================

    def _fill_statistics(
        self,
        ws,
        dashboard,
        rows,
    ):
        """
        Fill statistics worksheet.
        """

        ws["A1"] = "Fitness Controller PRO"

        ws["A1"].font = ExcelStyle.TITLE_FONT

        ws["A3"] = "Показатель"
        ws["B3"] = "Значение"

        ws["A3"].fill = ExcelStyle.HEADER_FILL
        ws["B3"].fill = ExcelStyle.HEADER_FILL

        ws["A3"].font = ExcelStyle.HEADER_FONT
        ws["B3"].font = ExcelStyle.HEADER_FONT

        stats = [

            (
                "Всего пользователей",
                dashboard["total_users"],
            ),

            (
                "Прошли сегодня",
                dashboard["completed"],
            ),

            (
                "Не прошли",
                dashboard["not_completed"],
            ),

            (
                "Процент выполнения",
                f"{dashboard['percent']} %",
            ),

            (
                "Всего ответов",
                len(rows),
            ),

            (
                "Всего фотографий",
                dashboard["photos"],
            ),

            (
                "Средний вес",
                dashboard["average_weight"],
            ),

        ]

        current = 4

        for title, value in stats:

            ws.cell(
                row=current,
                column=1,
                value=title,
            )

            ws.cell(
                row=current,
                column=2,
                value=value,
            )

            current += 1

        ws.freeze_panes = "A4"

    # =====================================================
    # PHOTOS
    # =====================================================

    def _fill_photos(
    self,
    ws,
    photos,
):
    """
    Prepare photo worksheet.

    Images are inserted later.
    """

    headers = [
        "Пользователь",
        "Дата",
        "Вопрос",
        "Фото",
    ]

    for col, title in enumerate(
        headers,
        start=1,
    ):

        cell = ws.cell(
            row=1,
            column=col,
            value=title,
        )

        cell.fill = ExcelStyle.HEADER_FILL
        cell.font = ExcelStyle.HEADER_FONT
        cell.alignment = ExcelStyle.CENTER
        cell.border = ExcelStyle.BORDER

    row = 2

    for item in photos:

        ws.cell(
            row=row,
            column=1,
            value=item["user"],
        )

        ws.cell(
            row=row,
            column=2,
            value=item["date"],
        )

        ws.cell(
            row=row,
            column=3,
            value=item["question"],
        )

        ws.cell(
            row=row,
            column=4,
            value=item["telegram_file_id"],
        )

        row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # =====================================================
    # SAVE
    # =====================================================

    def _save(
        self,
        workbook,
        filepath,
    ):
        """
        Save workbook.
        """

        workbook.save(
            filepath,
        )
